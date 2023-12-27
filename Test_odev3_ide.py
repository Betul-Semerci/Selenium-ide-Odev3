from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
import pytest
import openpyxl
from constants import globalConstants as c


class Test_Sauce_Odev1:
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window() 

    def teardown_method(self):
        self.driver.quit()

    def getData1():
        excel = openpyxl.load_workbook(c.invalid_login_xlsx)
        sheet = excel["Sayfa1"] 
        data = []
        for i in range(2,5): 
            username = sheet.cell(i,1).value 
            password = sheet.cell(i,2).value 
            data.append((username,password))
        return data
    
    def getData2():
        excel = openpyxl.load_workbook(c.invalid_login_xlsx)
        sheet = excel["Sayfa1"] 
        rows = sheet.max_row
        data = []
        for i in range(5,rows+1): 
            username = sheet.cell(i,1).value 
            password = sheet.cell(i,2).value 
            data.append((username,password))
        return data
    
    @pytest.mark.parametrize("username,password",getData1()) 
    def test_invalid_login(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,c.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == c.USERNAME_PASSWORD_DONT_MATCH


    @pytest.mark.parametrize("username,password",getData2()) 
    def test_urun_ekle(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.ADDTOCART_XPATH)))
        addToCart.click()
        shoppingCartLink = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.SHOPPINGCART_LINK_XPATH)))
        shoppingCartLink.click()
        product = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.PRODUCT_XPATH)))
        mesagge = product.text
        print(f"Sepetteki urun adı: {mesagge}")
        continueShopping = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.CONTINUE_SHOPPING_XPATH)))
        continueShopping.click()
        remove = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REMOVE_XPATH)))
        assert remove.text == c.REMOVE

    @pytest.mark.parametrize("username,password",getData2())  
    def test_urun_inceleme(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        self.driver.execute_script("window.scrollTo(0,500)")
        itemName= WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.ITEM_NAME_XPATH )))
        itemName.click()
        productTitle = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.PRODUCT_TITLE_XPATH)))
        message = productTitle.text
        print(f"Urunun Adi: {message}")
        backButton = self.driver.find_element(By.XPATH,c.BACK_BUTTON_XPATH)
        backButton.click()
        assert message == c.MESSAGE_XPATH
